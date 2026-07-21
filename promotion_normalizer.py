"""Validation and export helpers for the promotion time-series template.

The compact Support_Rules table is the source of truth.  A daily support file
is materialized only when the user requests it, so large promotions do not
inflate the source data permanently.
"""

from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Iterable


PROMOTION_MASTER_COLUMNS = (
    "promotion_id",
    "promotion_name",
    "channel",
    "promotion_type",
    "notes",
)
SUPPORT_RULE_COLUMNS = (
    "support_rule_id",
    "promotion_id",
    "model_code",
    "start_date",
    "end_date",
    "support_per_unit",
    "currency",
)
DAILY_SUPPORT_COLUMNS = (
    "applied_date",
    "model_code",
    "promotion_id",
    "support_rule_id",
    "support_per_unit",
    "currency",
)
EXCEL_MAX_DATA_ROWS = 1_048_575


@dataclass(frozen=True)
class ValidationIssue:
    sheet: str
    row: int
    column: str
    message: str

    def display(self) -> str:
        return f"{self.sheet} row {self.row} · {self.column}: {self.message}"


@dataclass(frozen=True)
class PromotionTemplateData:
    master_rows: tuple[dict, ...]
    support_rules: tuple[dict, ...]
    estimated_daily_rows: int
    overlapping_rule_pairs: int


@dataclass(frozen=True)
class PromotionExportResult:
    master_path: Path
    rules_path: Path
    daily_path: Path
    daily_rows: int
    overlapping_rule_pairs: int


class TemplateValidationError(ValueError):
    def __init__(self, issues: Iterable[ValidationIssue]):
        self.issues = tuple(issues)
        super().__init__("The promotion template contains validation errors.")


def _clean_text(value) -> str:
    return "" if value is None else str(value).strip()


def _parse_iso_date(value) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return date.fromisoformat(_clean_text(value))
    except ValueError:
        return None


def _parse_amount(value) -> Decimal | None:
    text = _clean_text(value).replace("\u00a0", " ").replace(" ", "")
    if not text:
        return None
    # The template uses an invariant decimal point.  A comma-only amount is
    # accepted as a convenient input form, but ambiguous mixed punctuation is
    # rejected instead of silently changing an amount.
    if "," in text and "." not in text:
        text = text.replace(",", ".")
    if text.count(".") > 1 or "," in text:
        return None
    try:
        amount = Decimal(text)
    except InvalidOperation:
        return None
    return amount if amount >= 0 else None


def _normalised_master_row(record: dict) -> dict:
    return {column: _clean_text(record.get(column)) for column in PROMOTION_MASTER_COLUMNS}


def _normalised_rule_row(record: dict) -> dict:
    return {column: record.get(column) for column in SUPPORT_RULE_COLUMNS}


def validate_records(master_records: Iterable[dict], rule_records: Iterable[dict]) -> tuple[PromotionTemplateData | None, tuple[ValidationIssue, ...]]:
    """Validate compact promotion records without expanding them to daily rows."""
    issues: list[ValidationIssue] = []
    masters: list[dict] = []
    rules: list[dict] = []
    promotion_ids: set[str] = set()
    rule_ids: set[str] = set()

    for index, raw in enumerate(master_records, start=2):
        row = _normalised_master_row(raw)
        promotion_id = row["promotion_id"]
        if not promotion_id:
            issues.append(ValidationIssue("Promotion_Master", index, "promotion_id", "A promotion ID is required."))
        elif promotion_id in promotion_ids:
            issues.append(ValidationIssue("Promotion_Master", index, "promotion_id", "Promotion ID must be unique."))
        else:
            promotion_ids.add(promotion_id)
        if not row["promotion_name"]:
            issues.append(ValidationIssue("Promotion_Master", index, "promotion_name", "A promotion name is required."))
        masters.append(row)

    if not masters:
        issues.append(ValidationIssue("Promotion_Master", 2, "sheet", "Add at least one promotion."))

    for index, raw in enumerate(rule_records, start=2):
        raw_row = _normalised_rule_row(raw)
        rule_id = _clean_text(raw_row["support_rule_id"])
        promotion_id = _clean_text(raw_row["promotion_id"])
        model_code = _clean_text(raw_row["model_code"])
        currency = _clean_text(raw_row["currency"])
        start_date = _parse_iso_date(raw_row["start_date"])
        end_date = _parse_iso_date(raw_row["end_date"])
        amount = _parse_amount(raw_row["support_per_unit"])

        if not rule_id:
            issues.append(ValidationIssue("Support_Rules", index, "support_rule_id", "A support rule ID is required."))
        elif rule_id in rule_ids:
            issues.append(ValidationIssue("Support_Rules", index, "support_rule_id", "Support rule ID must be unique."))
        else:
            rule_ids.add(rule_id)
        if not promotion_id:
            issues.append(ValidationIssue("Support_Rules", index, "promotion_id", "A promotion ID is required."))
        elif promotion_id not in promotion_ids:
            issues.append(ValidationIssue("Support_Rules", index, "promotion_id", "Promotion ID is not in Promotion_Master."))
        if not model_code:
            issues.append(ValidationIssue("Support_Rules", index, "model_code", "A model code is required."))
        if start_date is None:
            issues.append(ValidationIssue("Support_Rules", index, "start_date", "Use YYYY-MM-DD."))
        if end_date is None:
            issues.append(ValidationIssue("Support_Rules", index, "end_date", "Use YYYY-MM-DD."))
        if start_date and end_date and end_date < start_date:
            issues.append(ValidationIssue("Support_Rules", index, "end_date", "End date must not be before start date."))
        if amount is None:
            issues.append(ValidationIssue("Support_Rules", index, "support_per_unit", "Use a non-negative number."))
        if not currency:
            issues.append(ValidationIssue("Support_Rules", index, "currency", "A currency is required."))

        rules.append({
            "support_rule_id": rule_id,
            "promotion_id": promotion_id,
            "model_code": model_code,
            "start_date": start_date,
            "end_date": end_date,
            "support_per_unit": amount,
            "currency": currency,
        })

    if not rules:
        issues.append(ValidationIssue("Support_Rules", 2, "sheet", "Add at least one support rule."))

    if issues:
        return None, tuple(issues)

    estimated_daily_rows = sum((rule["end_date"] - rule["start_date"]).days + 1 for rule in rules)
    overlap_pairs = _count_overlapping_rule_pairs(rules)
    return PromotionTemplateData(tuple(masters), tuple(rules), estimated_daily_rows, overlap_pairs), ()


def _count_overlapping_rule_pairs(rules: Iterable[dict]) -> int:
    grouped: dict[str, list[dict]] = {}
    for rule in rules:
        grouped.setdefault(rule["model_code"], []).append(rule)
    overlaps = 0
    for model_rules in grouped.values():
        ordered = sorted(model_rules, key=lambda rule: (rule["start_date"], rule["end_date"]))
        for index, rule in enumerate(ordered):
            for later_rule in ordered[index + 1:]:
                if later_rule["start_date"] > rule["end_date"]:
                    break
                overlaps += 1
    return overlaps


def _sheet_records(worksheet, required_columns: tuple[str, ...]) -> tuple[list[dict], tuple[ValidationIssue, ...]]:
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return [], (ValidationIssue(worksheet.title, 1, "sheet", "The sheet is empty."),)
    headers = [_clean_text(value) for value in rows[0]]
    missing = [column for column in required_columns if column not in headers]
    if missing:
        return [], tuple(ValidationIssue(worksheet.title, 1, column, "Required column is missing.") for column in missing)

    records = []
    for values in rows[1:]:
        if not any(value not in (None, "") for value in values):
            continue
        records.append(dict(zip(headers, values)))
    return records, ()


def load_template(path: str | Path) -> tuple[PromotionTemplateData | None, tuple[ValidationIssue, ...]]:
    """Load and validate the two-sheet Excel template on demand."""
    import openpyxl

    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        issues: list[ValidationIssue] = []
        if "Promotion_Master" not in workbook.sheetnames:
            issues.append(ValidationIssue("Promotion_Master", 1, "sheet", "Required sheet is missing."))
            master_records = []
        else:
            master_records, sheet_issues = _sheet_records(workbook["Promotion_Master"], PROMOTION_MASTER_COLUMNS)
            issues.extend(sheet_issues)
        if "Support_Rules" not in workbook.sheetnames:
            issues.append(ValidationIssue("Support_Rules", 1, "sheet", "Required sheet is missing."))
            rule_records = []
        else:
            rule_records, sheet_issues = _sheet_records(workbook["Support_Rules"], SUPPORT_RULE_COLUMNS)
            issues.extend(sheet_issues)
        if issues:
            return None, tuple(issues)
        data, validation_issues = validate_records(master_records, rule_records)
        return data, validation_issues
    finally:
        workbook.close()


def preview_daily_rows(data: PromotionTemplateData, limit: int = 20) -> list[dict]:
    preview: list[dict] = []
    for rule in data.support_rules:
        current = rule["start_date"]
        while current <= rule["end_date"] and len(preview) < limit:
            preview.append(_daily_record(rule, current))
            current += timedelta(days=1)
        if len(preview) >= limit:
            break
    return preview


def _daily_record(rule: dict, applied_date: date) -> dict:
    return {
        "applied_date": applied_date.isoformat(),
        "model_code": rule["model_code"],
        "promotion_id": rule["promotion_id"],
        "support_rule_id": rule["support_rule_id"],
        "support_per_unit": format(rule["support_per_unit"], "f"),
        "currency": rule["currency"],
    }


def _next_output_path(directory: Path, basename: str, extension: str, stamp: str) -> Path:
    candidate = directory / f"{basename}_{stamp}{extension}"
    suffix = 2
    while candidate.exists():
        candidate = directory / f"{basename}_{stamp}_{suffix:02d}{extension}"
        suffix += 1
    return candidate


def _write_compact_csv(path: Path, columns: tuple[str, ...], rows: Iterable[dict]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns)
        writer.writeheader()
        for row in rows:
            writer.writerow({
                column: row[column].isoformat() if isinstance(row.get(column), date) else row.get(column, "")
                for column in columns
            })


def _iter_daily_records(rules: Iterable[dict]):
    for rule in rules:
        current = rule["start_date"]
        while current <= rule["end_date"]:
            yield _daily_record(rule, current)
            current += timedelta(days=1)


def _write_daily_csv(path: Path, rules: Iterable[dict]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=DAILY_SUPPORT_COLUMNS)
        writer.writeheader()
        writer.writerows(_iter_daily_records(rules))


def _write_daily_excel(path: Path, rules: Iterable[dict]) -> None:
    import openpyxl

    workbook = openpyxl.Workbook(write_only=True)
    worksheet = workbook.create_sheet("Daily_Support")
    worksheet.append(list(DAILY_SUPPORT_COLUMNS))
    for row in _iter_daily_records(rules):
        worksheet.append([row[column] for column in DAILY_SUPPORT_COLUMNS])
    workbook.save(path)


def export_normalized(data: PromotionTemplateData, source_path: str | Path, daily_format: str = "CSV", timestamp: datetime | None = None) -> PromotionExportResult:
    """Write compact source tables plus one analysis-ready daily support file."""
    source = Path(source_path)
    directory = source.parent
    stamp = (timestamp or datetime.now()).strftime("%Y%m%d_%H%M")
    master_path = _next_output_path(directory, "promotion_master", ".csv", stamp)
    rules_path = _next_output_path(directory, "promotion_support_rules", ".csv", stamp)
    if daily_format == "Excel (.xlsx)":
        if data.estimated_daily_rows > EXCEL_MAX_DATA_ROWS:
            raise ValueError("The daily result exceeds Excel's row limit. Select CSV output.")
        daily_path = _next_output_path(directory, "promotion_daily_support", ".xlsx", stamp)
    else:
        daily_path = _next_output_path(directory, "promotion_daily_support", ".csv", stamp)

    _write_compact_csv(master_path, PROMOTION_MASTER_COLUMNS, data.master_rows)
    _write_compact_csv(rules_path, SUPPORT_RULE_COLUMNS, data.support_rules)
    if daily_format == "Excel (.xlsx)":
        _write_daily_excel(daily_path, data.support_rules)
    else:
        _write_daily_csv(daily_path, data.support_rules)

    return PromotionExportResult(
        master_path=master_path,
        rules_path=rules_path,
        daily_path=daily_path,
        daily_rows=data.estimated_daily_rows,
        overlapping_rule_pairs=data.overlapping_rule_pairs,
    )
